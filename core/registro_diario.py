import os
import sys
import json
from datetime import datetime

from openpyxl import Workbook, load_workbook
from core import local_now


def _app_dir() -> str:
    """
    Carpeta base de la app:
    - En .exe (PyInstaller): carpeta donde está el .exe
    - En .py: raíz del proyecto (un nivel arriba de /core)
    """
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def _desktop_dir() -> str:
    """
    Devuelve el Escritorio real del usuario si existe, si no, HOME.
    """
    home = os.path.expanduser("~")
    desktop = os.path.join(home, "Desktop")
    if os.path.isdir(desktop):
        return desktop
    # Alternativa común en Windows en español (a veces no existe como carpeta física)
    desktop_alt = os.path.join(home, "Escritorio")
    if os.path.isdir(desktop_alt):
        return desktop_alt
    return home


def _base_dir() -> str:
    """
    Prioridades:
    1) D:\\Desktop si existe (tu estándar actual)
    2) Desktop real del usuario
    3) Carpeta del .exe / app
    """
    configured = os.environ.get("HELP_APP_DATA_DIR", "").strip()
    if configured:
        os.makedirs(configured, exist_ok=True)
        return os.path.abspath(configured)

    d_desktop = r"D:\Desktop"
    if os.path.isdir(d_desktop):
        return d_desktop

    desk = _desktop_dir()
    if os.path.isdir(desk):
        return desk

    return _app_dir()


# 📂 Carpeta base para tareas diarias (portable)
BASE_DIR = _base_dir()
CARPETA_TAREAS_DIARIAS = os.path.join(BASE_DIR, "HELP_APP_TAREAS_DIARIAS")


def get_carpeta_diaria():
    """
    Crea (si no existe) la carpeta general 'HELP_APP_TAREAS_DIARIAS'
    y, dentro, una subcarpeta con la fecha actual: YYYY_MM_DD.
    Devuelve la ruta completa de la carpeta del día.
    """
    os.makedirs(CARPETA_TAREAS_DIARIAS, exist_ok=True)

    hoy = local_now().strftime("%Y_%m_%d")
    carpeta_hoy = os.path.join(CARPETA_TAREAS_DIARIAS, hoy)
    os.makedirs(carpeta_hoy, exist_ok=True)

    print(f"Carpeta diaria lista: {carpeta_hoy}")
    return carpeta_hoy


def consolidar_mes_xlsx(mes=None):
    """
    Agrupa todas las carpetas YYYY_MM_DD del mes indicado y genera
    CARPETA_TAREAS_DIARIAS/YYYY_MM/tareas_mes.xlsx con una hoja 'Tareas'
    y una hoja 'checklist' consolidadas.

    mes: "YYYY_MM" o None = mes actual.
    Devuelve la ruta del archivo generado.
    """
    if mes is None:
        mes = local_now().strftime("%Y_%m")

    os.makedirs(CARPETA_TAREAS_DIARIAS, exist_ok=True)

    # Carpeta destino mensual
    carpeta_mes = os.path.join(CARPETA_TAREAS_DIARIAS, mes)
    os.makedirs(carpeta_mes, exist_ok=True)

    # Recolectar todas las carpetas del mes
    tareas_headers = None
    tareas_rows = []
    checklist_headers = None
    checklist_rows = []

    try:
        nombres = sorted(os.listdir(CARPETA_TAREAS_DIARIAS))
    except OSError:
        nombres = []

    for nombre in nombres:
        # Solo directorios YYYY_MM_DD que correspondan al mes
        parts = nombre.split("_")
        if len(parts) != 3 or not all(p.isdigit() for p in parts):
            continue
        if f"{parts[0]}_{parts[1]}" != mes:
            continue

        ruta_xlsx = os.path.join(CARPETA_TAREAS_DIARIAS, nombre, "tareas.xlsx")
        if not os.path.exists(ruta_xlsx):
            continue

        try:
            wb = load_workbook(ruta_xlsx, data_only=True, read_only=True)
        except Exception:
            continue

        try:
            # Hoja Tareas
            ws_t = wb["Tareas"] if "Tareas" in wb.sheetnames else wb.active
            filas_t = list(ws_t.iter_rows(values_only=True))
            if filas_t:
                hdrs = ["" if v is None else str(v).strip() for v in filas_t[0]]
                if tareas_headers is None:
                    tareas_headers = hdrs
                for fila in filas_t[1:]:
                    row = ["" if v is None else str(v) for v in fila]
                    if any(c != "" for c in row):
                        tareas_rows.append(row)

            # Hoja checklist
            if "checklist" in wb.sheetnames:
                ws_c = wb["checklist"]
                filas_c = list(ws_c.iter_rows(values_only=True))
                if filas_c:
                    hdrs_c = ["" if v is None else str(v).strip() for v in filas_c[0]]
                    if checklist_headers is None:
                        checklist_headers = hdrs_c
                    for fila in filas_c[1:]:
                        row = ["" if v is None else str(v) for v in fila]
                        if any(c != "" for c in row):
                            checklist_rows.append(row)
        finally:
            wb.close()

    # Escribir Excel mensual
    from openpyxl import Workbook as _Workbook
    wb_out = _Workbook()
    ws_out = wb_out.active
    ws_out.title = "Tareas"

    if tareas_headers:
        ws_out.append(tareas_headers)
    for row in tareas_rows:
        ws_out.append(row)

    if checklist_headers:
        ws_chk = wb_out.create_sheet(title="checklist")
        ws_chk.append(checklist_headers)
        for row in checklist_rows:
            ws_chk.append(row)

    ruta_salida = os.path.join(carpeta_mes, "tareas_mes.xlsx")
    wb_out.save(ruta_salida)
    print(f"[CONSOLIDAR_MES] Generado: {ruta_salida} ({len(tareas_rows)} tareas)")
    return ruta_salida


HEADERS_TAREAS = [
    "fecha_elaboracion",
    "OP No.",
    "descripcion",
    "proceso",
    "usuario",
    "estado",
    "cantidad",
    "inicio",
    "fin",
    "tiempo_tarea",
]

CHECKLIST_EXCLUDE_KEYS = {
    "id",
    "modulo",
    "tarea_id",
    "usuario",
    "op_no",
    "proceso",
    "descripcion",
    "estado",
    "titulo",
    "fecha_elaboracion",
    "fecha_checklist",
}

CHECKLIST_BASE_HEADERS = ["op", "tarea"]


def _ensure_headers(ws, headers):
    """
    Asegura que existan estos encabezados en la fila 1.
    - No inserta columnas (para no desplazar datos).
    - Si faltan encabezados, los agrega al final.
    """
    if ws.max_row < 1:
        ws.append(headers)
        return

    existentes = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    existentes_norm = [str(x).strip() if x is not None else "" for x in existentes]

    if all(x == "" for x in existentes_norm):
        for idx, h in enumerate(headers, start=1):
            ws.cell(row=1, column=idx).value = h
        return

    for h in headers:
        if h not in existentes_norm:
            ws.cell(row=1, column=ws.max_column + 1).value = h
            existentes_norm.append(h)


def _headers_row(ws):
    if ws.max_row < 1:
        return []
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        headers.append("" if v is None else str(v).strip())
    while headers and headers[-1] == "":
        headers.pop()
    return headers


def _normalizar_dict_checklist(data):
    if data in (None, "", {}):
        return {}

    raw = data
    if isinstance(raw, str):
        s = raw.strip()
        if not s:
            return {}
        try:
            raw = json.loads(s)
        except Exception:
            return {}

    if not isinstance(raw, dict):
        return {}

    out = {}
    for k, v in raw.items():
        key = str(k).strip()
        if not key:
            continue
        if key.lower() in CHECKLIST_EXCLUDE_KEYS:
            continue
        out[key] = "" if v is None else str(v)
    return out


def _migrar_hoja_checklist_legacy(ws):
    """
    Migra formato antiguo (columna checklist_json) a columnas por campo
    del checklist, conservando las filas ya registradas.
    """
    headers = _headers_row(ws)
    if not headers or "checklist_json" not in headers:
        return

    idx_json = headers.index("checklist_json") + 1
    idx_op = None
    idx_tarea = None
    for i, h in enumerate(headers, start=1):
        hl = (h or "").strip().lower()
        if idx_op is None and hl in {"op", "op no.", "op no", "op_no", "titulo"}:
            idx_op = i
        if idx_tarea is None and hl in {"tarea", "proceso"}:
            idx_tarea = i

    rows_data = []
    keys = list(CHECKLIST_BASE_HEADERS)

    for r in range(2, ws.max_row + 1):
        raw = ws.cell(row=r, column=idx_json).value
        d = {
            "op": (ws.cell(row=r, column=idx_op).value if idx_op else "") or "",
            "tarea": (ws.cell(row=r, column=idx_tarea).value if idx_tarea else "") or "",
        }
        d_check = _normalizar_dict_checklist(raw)
        d.update(d_check)
        if not d_check and not d.get("op") and not d.get("tarea"):
            continue
        rows_data.append(d)
        for k in d_check.keys():
            if k not in keys:
                keys.append(k)

    total = ws.max_row
    if total > 0:
        ws.delete_rows(1, total)

    if not keys:
        return

    ws.append(keys)
    for d in rows_data:
        ws.append([d.get(k, "") for k in keys])


def _normalizar_hoja_checklist(ws):
    """
    Asegura que 'op' y 'tarea' existan y queden al inicio de la hoja checklist.
    Si la hoja ya tiene datos, reordena columnas preservando filas.
    """
    headers = _headers_row(ws)
    if not headers:
        ws.append(list(CHECKLIST_BASE_HEADERS))
        return

    target_headers = list(CHECKLIST_BASE_HEADERS)
    for h in headers:
        hh = (h or "").strip()
        if not hh or hh in target_headers:
            continue
        target_headers.append(hh)

    if headers == target_headers:
        return

    rows_dict = []
    for r in range(2, ws.max_row + 1):
        d = {}
        has_val = False
        for c, h in enumerate(headers, start=1):
            key = (h or "").strip()
            if not key:
                continue
            val = ws.cell(row=r, column=c).value
            if val not in (None, ""):
                has_val = True
            d[key] = val
        if has_val:
            rows_dict.append(d)

    total = ws.max_row
    if total > 0:
        ws.delete_rows(1, total)

    ws.append(target_headers)
    for d in rows_dict:
        ws.append([d.get(h, "") for h in target_headers])


def _abrir_o_crear_archivo_diario():
    carpeta = get_carpeta_diaria()
    ruta = os.path.join(carpeta, "tareas.xlsx")

    if os.path.exists(ruta):
        wb = load_workbook(ruta)
    else:
        wb = Workbook()

    # ✅ Si existe hoja OPs, la eliminamos (ya no la usamos)
    if "OPs" in wb.sheetnames:
        try:
            wb.remove(wb["OPs"])
        except Exception:
            # Best-effort: si por alguna razón falla, no frenamos el sistema
            pass

    # Hoja detalle (una fila por tarea)
    if "Tareas" in wb.sheetnames:
        ws_tareas = wb["Tareas"]
    else:
        ws_tareas = wb.active
        ws_tareas.title = "Tareas"
    _ensure_headers(ws_tareas, HEADERS_TAREAS)

    # Hoja de checklist (una fila por checklist asociado a tarea finalizada)
    if "checklist" in wb.sheetnames:
        ws_checklist = wb["checklist"]
    else:
        ws_checklist = wb.create_sheet(title="checklist")
    _migrar_hoja_checklist_legacy(ws_checklist)
    _normalizar_hoja_checklist(ws_checklist)

    wb.save(ruta)
    return wb, ws_tareas, ws_checklist, ruta


def _calcular_tiempo_tarea(inicio: str, fin: str) -> str:
    if not inicio or not fin:
        return ""
    try:
        fmt = "%Y-%m-%d %H:%M:%S"
        dt_i = datetime.strptime(inicio, fmt)
        dt_f = datetime.strptime(fin, fmt)
        seg = int((dt_f - dt_i).total_seconds())
        if seg < 0:
            return ""
        h = seg // 3600
        m = (seg % 3600) // 60
        s = seg % 60
        return f"{h:02d}:{m:02d}:{s:02d}"
    except Exception:
        return ""


def _calcular_fecha_elaboracion(fin: str) -> str:
    if fin:
        try:
            fmt = "%Y-%m-%d %H:%M:%S"
            return datetime.strptime(fin, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    return local_now().strftime("%Y-%m-%d")


# ==========================================================
# ✅ Formateo HH:MM (solo para guardar/mostrar en Excel)
# ==========================================================
def _a_hhmm(value: str) -> str:
    """
    Convierte:
      - 'YYYY-MM-DD HH:MM:SS' -> 'HH:MM'
      - 'HH:MM:SS' -> 'HH:MM'
      - 'HH:MM' -> 'HH:MM'
    Si no puede, devuelve el valor original.
    """
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""

    # Si viene con fecha, tomar parte de hora
    if " " in s:
        s = s.split(" ")[-1]
    if "T" in s:
        s = s.split("T")[-1]

    parts = s.split(":")
    if len(parts) >= 2:
        try:
            h = int(parts[0])
            m = int(parts[1])
            return f"{h:02d}:{m:02d}"
        except Exception:
            return s
    return s


def registrar_tarea_diaria(
    titulo="",        # OP No.
    descripcion="",
    proceso="",
    usuario="",
    estado="",
    cantidad=0,
    inicio="",
    fin="",
    checklist_modulo="",
    checklist_tarea_id=None,
    checklist_fecha="",
    checklist_data=None,
):
    """
    - Hoja 'Tareas': APPEND (una fila por cada tarea finalizada).

    ✅ Cambios:
    - Se elimina la hoja 'OPs' del Excel (ya no se usa).
    - 'inicio', 'fin', 'tiempo_tarea' se guardan como HH:MM en Excel.
    """
    wb, ws_tareas, ws_checklist, ruta = _abrir_o_crear_archivo_diario()

    # Se calcula con los timestamps completos (no cambia lógica)
    tiempo = _calcular_tiempo_tarea(inicio, fin)
    fecha_elab = _calcular_fecha_elaboracion(fin)

    # ✅ Solo para Excel (formato HH:MM)
    inicio_hhmm = _a_hhmm(inicio)
    fin_hhmm = _a_hhmm(fin)
    tiempo_hhmm = _a_hhmm(tiempo)

    op_no = (titulo or "").strip()

    ws_tareas.append([
        fecha_elab,
        op_no,
        (descripcion or ""),
        (proceso or ""),
        (usuario or ""),
        (estado or ""),
        int(cantidad) if cantidad is not None else "",
        (inicio_hhmm or ""),
        (fin_hhmm or ""),
        (tiempo_hhmm or ""),
    ])

    checklist_dict = _normalizar_dict_checklist(checklist_data)
    if checklist_dict:
        checklist_dict = {
            "op": op_no,
            "tarea": (proceso or "").strip(),
            **checklist_dict,
        }
        current_headers = _headers_row(ws_checklist)
        if not current_headers:
            ws_checklist.append(list(CHECKLIST_BASE_HEADERS))
            current_headers = _headers_row(ws_checklist)

        _ensure_headers(ws_checklist, list(checklist_dict.keys()))
        _normalizar_hoja_checklist(ws_checklist)
        current_headers = _headers_row(ws_checklist)

        ws_checklist.append([checklist_dict.get(h, "") for h in current_headers])

    wb.save(ruta)
    print(f"[REGISTRO_DIARIO] Actualizado: {ruta}")
