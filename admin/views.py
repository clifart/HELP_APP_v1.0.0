from flask import render_template, request, redirect, url_for, session, flash, current_app
from . import admin_bp
from core.db import get_db_connection, ensure_tareas_columns
from core.historial_central import (
    HISTORIAL_CENTRAL_DIR,
    limpiar_historial_con_respaldo,
    respaldar_historial_en_central,
)
from core.registro_diario import registrar_tarea_diaria, CARPETA_TAREAS_DIARIAS  # ðŸ‘ˆ registro diario
from core.exportar_excel import exportar_excel_automatico  # ðŸ‘ˆ
import sqlite3
import os
import sys
import subprocess
from flask import current_app
from core import now_iso
from flask import send_from_directory, abort
from datetime import datetime




# ðŸ”‘ CLAVE ÃšNICA PARA TODO EL ADMIN
CLAVE_TECNICA = os.environ.get("HELP_APP_MASTER_KEY", "HELPAPP_2025")


def _solo_admin():
    return session.get("rol") == "admin"


# ðŸ‘‡ NUEVO: helper para saber si Modo TÃ©cnico estÃ¡ activo
def _modo_tecnico_activo():
    return session.get("clave_tecnica") == CLAVE_TECNICA


def _asegurar_cols_historial(cur):
    cur.execute("""
        CREATE TABLE IF NOT EXISTS historial_tareas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario TEXT,
            op_no TEXT,
            descripcion TEXT,
            tarea TEXT,
            cantidad INTEGER,
            hora_inicio TEXT,
            hora_finalizacion TEXT,
            tiempo_total TEXT,
            horas_extras TEXT,
            tiempo_pausado TEXT,
            cierre_automatico INTEGER DEFAULT 0
        )
    """)
    cur.execute("PRAGMA table_info(historial_tareas)")
    cols = [row[1] for row in cur.fetchall()]
    if "tiempo_total" not in cols:
        try:
            cur.execute("ALTER TABLE historial_tareas ADD COLUMN tiempo_total TEXT")
        except sqlite3.OperationalError:
            pass
    if "horas_extras" not in cols:
        try:
            cur.execute("ALTER TABLE historial_tareas ADD COLUMN horas_extras TEXT")
        except sqlite3.OperationalError:
            pass
    if "tiempo_pausado" not in cols:
        try:
            cur.execute("ALTER TABLE historial_tareas ADD COLUMN tiempo_pausado TEXT")
        except sqlite3.OperationalError:
            pass
    if "cierre_automatico" not in cols:
        try:
            cur.execute("ALTER TABLE historial_tareas ADD COLUMN cierre_automatico INTEGER DEFAULT 0")
        except sqlite3.OperationalError:
            pass


# ==========================================================
# PANEL ADMIN
# ==========================================================
@admin_bp.route("/")
def panel():
    if not _solo_admin():
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("login"))

    conn = get_db_connection()
    cur = conn.cursor()

    # ðŸ‘‡ Crear tabla tareas si no existe (por si la BD quedÃ³ limpia)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS tareas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            titulo TEXT,
            descripcion TEXT,
            proceso TEXT,
            cantidad INTEGER,
            asignado_a TEXT,
            inicio TEXT,
            fin TEXT,
            estado TEXT
        )
    """)

    # Ajustar columnas extra si las hubiera
    ensure_tareas_columns(conn)

    # ðŸ”¹ EstadÃ­sticas
    cur.execute("SELECT COUNT(*) FROM usuarios WHERE rol = 'usuario'")
    total_usuarios = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM usuarios WHERE rol = 'admin'")
    total_admins = cur.fetchone()[0]

    cur.execute("""
        SELECT COUNT(*)
        FROM tareas
        WHERE estado != 'Finalizado'
          AND (asignado_a IS NULL OR asignado_a = '')
    """)
    total_tareas = cur.fetchone()[0]

    # ðŸ”¹ Usuarios para la tabla
    usuarios = cur.execute(
        "SELECT * FROM usuarios ORDER BY id DESC"
    ).fetchall()

    # ðŸ”¹ OPs activas del ADMIN (catÃ¡logo)
    tareas = cur.execute("""
        SELECT
            t_admin.id,
            t_admin.titulo,
            t_admin.descripcion,
            t_admin.proceso,
            t_admin.estado,
            CASE
                WHEN EXISTS (
                    SELECT 1
                    FROM tareas t_user
                    WHERE t_user.titulo = t_admin.titulo
                      AND t_user.asignado_a IS NOT NULL
                      AND t_user.asignado_a != ''
                      AND t_user.proceso IS NOT NULL
                      AND TRIM(t_user.proceso) != ''
                )
                THEN 1 ELSE 0
            END AS tiene_tareas
        FROM tareas t_admin
        WHERE t_admin.estado != 'Finalizado'
          AND (t_admin.asignado_a IS NULL OR asignado_a = '')
        ORDER BY t_admin.id DESC
    """).fetchall()

    conn.close()

    # âœ… CAMBIO ÃšNICO: renderizar el panel V2 (nuevo look)
    return render_template(
        "admin_v2/panel.html",
        total_usuarios=total_usuarios,
        total_admins=total_admins,
        total_tareas=total_tareas,
        usuarios=usuarios,
        tareas=tareas
    )
# âœ… Ruta de prueba directa (por si el login redirige raro)
@admin_bp.route("/panel_v2")
def panel_v2():
    return panel()

# âœ… Panel viejo (IMG1) disponible como respaldo
@admin_bp.route("/panel_viejo")
def panel_viejo():
    if not _solo_admin():
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("login"))

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS tareas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            titulo TEXT,
            descripcion TEXT,
            proceso TEXT,
            cantidad INTEGER,
            asignado_a TEXT,
            inicio TEXT,
            fin TEXT,
            estado TEXT
        )
    """)
    ensure_tareas_columns(conn)

    cur.execute("SELECT COUNT(*) FROM usuarios WHERE rol = 'usuario'")
    total_usuarios = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM usuarios WHERE rol = 'admin'")
    total_admins = cur.fetchone()[0]

    cur.execute("""
        SELECT COUNT(*)
        FROM tareas
        WHERE estado != 'Finalizado'
          AND (asignado_a IS NULL OR asignado_a = '')
    """)
    total_tareas = cur.fetchone()[0]

    usuarios = cur.execute("SELECT * FROM usuarios ORDER BY id DESC").fetchall()

    tareas = cur.execute("""
        SELECT
            t_admin.id,
            t_admin.titulo,
            t_admin.descripcion,
            t_admin.proceso,
            t_admin.estado,
            CASE
                WHEN EXISTS (
                    SELECT 1
                    FROM tareas t_user
                    WHERE t_user.titulo = t_admin.titulo
                      AND t_user.asignado_a IS NOT NULL
                      AND t_user.asignado_a != ''
                      AND t_user.proceso IS NOT NULL
                      AND TRIM(t_user.proceso) != ''
                )
                THEN 1 ELSE 0
            END AS tiene_tareas
        FROM tareas t_admin
        WHERE t_admin.estado != 'Finalizado'
          AND (t_admin.asignado_a IS NULL OR asignado_a = '')
        ORDER BY t_admin.id DESC
    """).fetchall()

    conn.close()

    return render_template(
        "admin/panel.html",
        total_usuarios=total_usuarios,
        total_admins=total_admins,
        total_tareas=total_tareas,
        usuarios=usuarios,
        tareas=tareas
    )

# ==========================================================
# ABRIR CARPETA HELP_APP_TAREAS_DIARIAS (BOTÃ“N EN ADMIN/PANEL)
# ==========================================================
def _abrir_carpeta_sistema(ruta: str) -> bool:
    """
    Abre una carpeta y (en Windows) intenta traer el Explorador al primer plano.
    """
    try:
        if not ruta or not os.path.exists(ruta):
            return False

        if os.name == "nt":
            # 1) Abrir la carpeta
            subprocess.Popen(["explorer", ruta], shell=False)

            # 2) Best-effort: enfocar esa ventana del Explorador
            ps = r"""
$ErrorActionPreference='SilentlyContinue'
$target = (Resolve-Path '%RUTA%').Path
$url = 'file:///' + ($target -replace '\\','/')

Add-Type @"
using System;
using System.Runtime.InteropServices;
public class Win32 {
  [DllImport("user32.dll")] public static extern bool SetForegroundWindow(IntPtr hWnd);
  [DllImport("user32.dll")] public static extern bool ShowWindowAsync(IntPtr hWnd, int nCmdShow);
}
"@ | Out-Null

$shell = New-Object -ComObject Shell.Application
for ($i=0; $i -lt 20; $i++) {
  Start-Sleep -Milliseconds 150
  foreach ($w in $shell.Windows()) {
    if ($w -and $w.LocationURL -eq $url) {
      [Win32]::ShowWindowAsync([IntPtr]$w.HWND, 9) | Out-Null
      [Win32]::SetForegroundWindow([IntPtr]$w.HWND) | Out-Null
      exit 0
    }
  }
}
exit 0
"""
            ps = ps.replace("%RUTA%", ruta.replace("'", "''"))

            subprocess.Popen(
                ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps],
                shell=False,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL
            )
            return True

        if sys.platform == "darwin":
            subprocess.Popen(["open", ruta])
            return True

        subprocess.Popen(["xdg-open", ruta])
        return True

    except Exception:
        current_app.logger.exception("No se pudo abrir/enfocar la carpeta: %s", ruta)
        return False


from flask import redirect, url_for

@admin_bp.route("/abrir_tareas_diarias", methods=["POST"])
def abrir_tareas_diarias():
    # En servidor/web: redirige a la vista web de la carpeta diaria
    return redirect(url_for("admin.carpeta_diaria"))

import os
from flask import render_template

@admin_bp.route("/tareas_diarias")
def tareas_diarias():
    from flask import request
    import os
    import datetime

    # âœ… Misma base que usa carpeta_diaria (portable: local o PythonAnywhere)
    base = _carpeta_diaria_base()
    os.makedirs(base, exist_ok=True)

    # Recolectar directorios
    candidatos = []
    for name in os.listdir(base):
        full = os.path.join(base, name)
        if os.path.isdir(full):
            # Intentar parsear YYYY_MM_DD
            dt = None
            try:
                parts = name.split("_")
                if len(parts) == 3 and all(p.isdigit() for p in parts):
                    y, m, d = map(int, parts)
                    dt = datetime.datetime(y, m, d)
            except Exception:
                dt = None

            if dt is None:
                # Fallback: fecha de modificaciÃ³n
                try:
                    dt = datetime.datetime.fromtimestamp(os.path.getmtime(full))
                except Exception:
                    dt = datetime.datetime.fromtimestamp(0)

            candidatos.append((dt, name))

    # Ordenar por mÃ¡s reciente y tomar 3
    candidatos.sort(key=lambda x: x[0], reverse=True)
    dias = [name for _, name in candidatos]

    # âœ… DEBUG: si pones ?debug=1 te muestra exactamente quÃ© estÃ¡ usando
    if request.args.get("debug") == "1":
        return (
            "DEBUG tareas_diarias\n"
            f"FILE={__file__}\n"
            f"BASE={base}\n"
            f"TOTAL_DIRS={len(candidatos)}\n"
            f"SHOWING={dias}\n"
        ), 200, {"Content-Type": "text/plain; charset=utf-8"}

    return render_template("admin/tareas_diarias.html", base=base, dias=dias)






# ==========================================================
# USUARIOS
# ==========================================================
@admin_bp.route('/agregar_usuario', methods=['POST'])
def agregar_usuario():
    if not _solo_admin():
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for('login'))

    # âœ… SOLO el nombre en MAYÃšSCULA
    nombre = request.form.get('nombre', '').strip().upper()

    celular = request.form.get('celular', '').strip()
    rol = request.form.get('rol', '').strip().lower()
    rol = 'admin' if 'admin' in rol else 'usuario'

    if not nombre:
        flash("âš ï¸ El nombre es obligatorio.", "warning")
        return redirect(url_for('admin.panel'))

    # ðŸ‘‡ NUEVO: crear ADMIN solo desde Modo TÃ©cnico
    if rol == 'admin' and not _modo_tecnico_activo():
        flash("âš ï¸ SÃ³lo puedes crear administradores desde Modo TÃ©cnico.", "danger")
        return redirect(url_for('admin.panel'))

    conn = get_db_connection()
    cur = conn.cursor()

    # âœ… Evita duplicados ignorando mayÃºsc/minÃºsc
    cur.execute("SELECT 1 FROM usuarios WHERE UPPER(TRIM(nombre)) = ?", (nombre,))
    if cur.fetchone():
        conn.close()
        flash("âš ï¸ El usuario ya existe.", "danger")
        return redirect(url_for('admin.panel'))

    cur.execute(
        "INSERT INTO usuarios (nombre, celular, contrasena, rol) VALUES (?, ?, ?, ?)",
        (nombre, celular, None, rol)
    )
    conn.commit()
    conn.close()

    flash(f"âœ… Usuario '{nombre}' agregado. CrearÃ¡ su contraseÃ±a en el primer ingreso.", "success")
    return redirect(url_for('admin.panel'))



@admin_bp.route("/eliminar_usuario/<int:id>", methods=["POST", "GET"])
def eliminar_usuario(id):
    if not _solo_admin():
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("login"))

    conn = get_db_connection()
    cur = conn.cursor()
    ensure_tareas_columns(conn)

    # ðŸ”¹ 1) Buscar el usuario (incluyendo rol)
    usuario = cur.execute(
        "SELECT id, nombre, rol FROM usuarios WHERE id = ?",
        (id,)
    ).fetchone()

    if not usuario:
        conn.close()
        flash("El usuario no existe.", "warning")
        return redirect(url_for("admin.panel"))

    nombre = usuario["nombre"]
    rol_usuario = usuario["rol"]

    # ðŸ‘‡ NUEVO: eliminar ADMIN solo desde Modo TÃ©cnico
    if rol_usuario == 'admin' and not _modo_tecnico_activo():
        conn.close()
        flash("âš ï¸ SÃ³lo puedes eliminar administradores desde Modo TÃ©cnico.", "danger")
        return redirect(url_for("admin.panel"))

    # ðŸ”¹ 2) Revisar si tiene tareas en curso
    cur.execute("""
        SELECT DISTINCT titulo
        FROM tareas
        WHERE asignado_a = ? AND estado != 'Finalizado'
    """, (nombre,))
    pendientes = cur.fetchall()

    if pendientes:
        ops = ", ".join([p["titulo"] for p in pendientes])
        flash(
            f"âš ï¸ No puedes eliminar al usuario {nombre} porque tiene tareas en curso "
            f"con las OP: {ops}. PÃ­dele que finalice sus tareas antes de eliminarlo.",
            "warning"
        )
        conn.close()
        return redirect(url_for("admin.panel"))

    # ðŸ”¹ 3) Si no tiene tareas pendientes, se elimina normalmente
    cur.execute("DELETE FROM usuarios WHERE id = ?", (id,))
    conn.commit()
    conn.close()

    flash(f"âœ… Usuario {nombre} eliminado correctamente.", "success")
    return redirect(url_for("admin.panel"))


# ==========================================================
# TAREAS / OP DEL ADMIN
# ==========================================================
@admin_bp.route('/agregar_op', methods=['POST'])
def agregar_op():
    if not _solo_admin():
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for('login'))

    # âœ… SOLO OP y DescripciÃ³n en MAYÃšSCULA
    titulo = request.form.get('titulo', '').strip().upper()       # OP No.
    descripcion = request.form.get('descripcion', '').strip().upper()
    proceso = request.form.get('proceso', '').strip()

    if not titulo or not descripcion:
        flash("âš ï¸ OP y DescripciÃ³n son obligatorias.", "warning")
        return redirect(url_for('admin.panel'))

    conn = get_db_connection()
    ensure_tareas_columns(conn)
    cur = conn.cursor()

    # 1ï¸âƒ£ Â¿Ya existe ESA OP ACTIVA en tareas? (ignorar mayÃºsc/minÃºsc)
    existe_activa = cur.execute("""
        SELECT 1
        FROM tareas
        WHERE UPPER(TRIM(titulo)) = ?
          AND estado != 'Finalizado'
        LIMIT 1
    """, (titulo,)).fetchone()

    if existe_activa:
        flash("âš ï¸ Ya existe una OP activa con ese nÃºmero. Usa otro.", "warning")
        conn.close()
        return redirect(url_for('admin.panel'))

    # 2ï¸âƒ£ Â¿Ya existe en el HISTORIAL? (OP marcada como Producto terminado)
    #    Si estÃ¡ en historial, no se puede volver a usar nunca.
    try:
        existe_historial = cur.execute("""
            SELECT 1
            FROM historial_tareas
            WHERE UPPER(TRIM(op_no)) = ?
            LIMIT 1
        """, (titulo,)).fetchone()
    except sqlite3.OperationalError:
        # Por si aÃºn no existe la tabla de historial en instalaciones nuevas
        existe_historial = None

    if existe_historial:
        flash("âš ï¸ Esta OP ya fue marcada como 'Producto terminado' "
              "y no se puede volver a usar.", "warning")
        conn.close()
        return redirect(url_for('admin.panel'))

    # 3ï¸âƒ£ Crear la nueva OP en tareas âœ… (SIN creado_en)
    cur.execute("""
        INSERT INTO tareas (titulo, descripcion, proceso, estado)
        VALUES (?, ?, ?, 'Pendiente')
    """, (titulo, descripcion, proceso))

    conn.commit()
    conn.close()
    flash("âœ… OP agregada correctamente.", "success")
    return redirect(url_for('admin.panel'))



@admin_bp.route('/agregar_tarea', methods=['POST'])
def agregar_tarea():
    # Alias para el formulario que llama a admin.agregar_tarea
    return agregar_op()


# ==========================================================
# FINALIZAR / ELIMINAR OP DEL ADMIN (BOTÃ“N EN PANEL)
# ==========================================================
@admin_bp.route('/finalizar_tarea/<int:id>', methods=['POST'])
def finalizar_tarea(id):
    """
    LÃ³gica del botÃ³n del admin:

    - Si hay tareas EN CURSO de esa OP (usuarios trabajando) -> NO hace nada, solo mensaje.
    - Si NO hay tareas de usuarios (nadie la tomÃ³ nunca) -> elimina la OP sin mandarla a historial.
    - Si hubo tareas de usuarios (aunque ya no estÃ©n en curso) -> marca la OP como Finalizada,
      toma la cantidad y el proceso de la ÃšLTIMA tarea finalizada del usuario
      y la registra en el diario (Excel).
    """
    if not _solo_admin():
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for('login'))

    conn = get_db_connection()
    # Si quieres usar acceso por nombre de columna, puedes activar row_factory:
    # conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    ensure_tareas_columns(conn)

    # 1ï¸âƒ£ Buscar la fila maestra del admin
    fila_admin = cur.execute(
        "SELECT titulo, descripcion, proceso FROM tareas WHERE id=?",
        (id,)
    ).fetchone()

    if not fila_admin:
        conn.close()
        flash("âš ï¸ No se encontrÃ³ la OP seleccionada.", "warning")
        return redirect(url_for('admin.panel'))

    # Compatibilidad por si no es sqlite.Row
    if hasattr(fila_admin, "keys"):
        op_titulo = fila_admin["titulo"]
        op_desc = fila_admin["descripcion"]
        op_proc = fila_admin["proceso"]
    else:
        op_titulo = fila_admin[0]
        op_desc = fila_admin[1]
        op_proc = fila_admin[2]

    # 2ï¸âƒ£ Verificar si hay tareas EN CURSO para esa OP tomadas por usuarios
    filas_en_curso = cur.execute("""
        SELECT asignado_a
        FROM tareas
        WHERE titulo = ?
          AND estado = 'En curso'
          AND asignado_a IS NOT NULL
          AND TRIM(asignado_a) != ''
    """, (op_titulo,)).fetchall()

    if filas_en_curso:
        # Hay usuarios trabajÃ¡ndola ahora mismo -> NO se puede finalizar
        nombres = set()
        for fila in filas_en_curso:
            if hasattr(fila, "keys"):
                nombre = fila["asignado_a"]
            else:
                nombre = fila[0]
            if nombre and str(nombre).strip():
                nombres.add(str(nombre).strip())

        lista_nombres = sorted(nombres)

        if lista_nombres:
            if len(lista_nombres) == 1:
                texto_usuarios = f"{lista_nombres[0]} tiene"
            elif len(lista_nombres) == 2:
                texto_usuarios = f"{lista_nombres[0]} y {lista_nombres[1]} tienen"
            else:
                texto_usuarios = (
                    ", ".join(lista_nombres[:-1]) +
                    f" y {lista_nombres[-1]} tienen"
                )
        else:
            texto_usuarios = "algÃºn usuario tiene"

        conn.close()
        flash(
            f"âš ï¸ No puedes finalizar la OP {op_titulo} "
            f"porque {texto_usuarios} tareas en curso con esa OP.",
            "warning"
        )
        return redirect(url_for('admin.panel'))

    # 3ï¸âƒ£ Verificar si ALGUNA VEZ tuvo tareas de usuarios (en cualquier estado)
    fila_uso = cur.execute("""
        SELECT 1
        FROM tareas
        WHERE titulo = ?
          AND asignado_a IS NOT NULL
          AND TRIM(asignado_a) != ''
        LIMIT 1
    """, (op_titulo,)).fetchone()

    # ðŸ§¹ CASO A: NUNCA FUE USADA POR NINGÃšN USUARIO -> ELIMINAR SIN HISTORIAL
    if not fila_uso:
        cur.execute("DELETE FROM tareas WHERE id = ?", (id,))
        conn.commit()
        conn.close()

        flash(
            f"â„¹ï¸ La OP {op_titulo} no tenÃ­a tareas registradas por ningÃºn usuario. "
            "Se eliminÃ³ sin enviarla al historial.",
            "info"
        )
        return redirect(url_for('admin.panel'))

    # ðŸ“’ CASO B: SÃ TUVO TAREAS DE USUARIOS (pero ninguna en curso)
    # ðŸ‘‰ 3.1 Tomar la ÃšLTIMA tarea del usuario desde historial_tareas
    usuario_excel = session.get("usuario", "admin")
    proceso_excel = op_proc
    cantidad_final = 0
    inicio_excel = ""
    fin_excel = ""

    try:
        fila_ultima_hist = cur.execute("""
            SELECT usuario, tarea, cantidad, hora_inicio, hora_finalizacion
            FROM historial_tareas
            WHERE op_no = ?
            ORDER BY datetime(hora_finalizacion) DESC, id DESC
            LIMIT 1
        """, (op_titulo,)).fetchone()
    except sqlite3.OperationalError:
        fila_ultima_hist = None

    if fila_ultima_hist:
        if hasattr(fila_ultima_hist, "keys"):
            usuario_excel = fila_ultima_hist["usuario"] or usuario_excel
            proceso_excel = fila_ultima_hist["tarea"] or proceso_excel
            cantidad_final = fila_ultima_hist["cantidad"] or 0
            inicio_excel = fila_ultima_hist["hora_inicio"] or ""
            fin_excel = fila_ultima_hist["hora_finalizacion"] or ""
        else:
            usuario_excel = fila_ultima_hist[0] or usuario_excel
            proceso_excel = fila_ultima_hist[1] or proceso_excel
            cantidad_final = fila_ultima_hist[2] or 0
            inicio_excel = fila_ultima_hist[3] or ""
            fin_excel = fila_ultima_hist[4] or ""
    else:
        # ðŸ” Respaldo: si por algÃºn motivo no hay historial, usamos la Ãºltima tarea finalizada en tareas
        fila_ultima_tareas = cur.execute("""
            SELECT cantidad, fin
            FROM tareas
            WHERE titulo = ?
              AND asignado_a IS NOT NULL
              AND TRIM(asignado_a) != ''
              AND estado = 'Finalizado'
            ORDER BY datetime(fin) DESC, id DESC
            LIMIT 1
        """, (op_titulo,)).fetchone()

        if fila_ultima_tareas:
            cantidad_final = fila_ultima_tareas[0] or 0
            fin_excel = fila_ultima_tareas[1] or ""

    # 3.2 Actualizar la OP maestra con estado finalizado y cantidad_final
    fin = now_iso()
    cur.execute("""
    UPDATE tareas
    SET estado='Finalizado', cantidad=?, fin=?
    WHERE id=?
    """, (cantidad_final, fin, id))

    conn.commit()
    conn.close()

    # ðŸ“ Registrar en el archivo diario (Excel)
    # registrar_tarea_diaria calcula internamente 'tiempo_tarea' a partir de inicio y fin
    try:
        registrar_tarea_diaria(
            titulo=op_titulo,
            descripcion=op_desc,
            proceso=proceso_excel,          # CORTE, IMPRESION, etc.
            usuario=usuario_excel,
            estado="Finalizado",
            cantidad=int(cantidad_final),
            inicio=inicio_excel,
            fin=fin_excel
        )
    except Exception as e:
        print(f"âš ï¸ No se pudo registrar la finalizaciÃ³n en el diario: {e}")

    flash("âœ… OP finalizada correctamente y registrado en el historial diario.", "success")
    return redirect(url_for('admin.panel'))



# ==========================================================
# HISTORIAL ADMIN â€” SOLO OP MAESTRAS FINALIZADAS DEL ADMIN
# ==========================================================
@admin_bp.route('/historial')
def historial():
    if not _solo_admin():
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("login"))

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    ensure_tareas_columns(conn)
    _asegurar_cols_historial(cur)

    # âœ… Solo OP maestras (del ADMIN) finalizadas:
    #    - estado = 'Finalizado'
    #    - asignado_a vacÃ­o (NULL o cadena vacÃ­a)
    # âœ… AdemÃ¡s: sacar el OPERARIO (primer registro en historial_tareas para esa OP)
    tareas = cur.execute("""
        SELECT
            t.id,
            t.titulo,
            t.descripcion,

            -- ðŸ‘‡ Operario que arrancÃ³ el proceso (ImpresiÃ³n en tu flujo)
            (
                SELECT h.usuario
                FROM historial_tareas h
                WHERE h.op_no = t.titulo
                ORDER BY h.id ASC       -- primero que registrÃ³ algo para esa OP
                LIMIT 1
            ) AS operario_impresion,

            t.cantidad,
            t.fin,
            EXISTS(
                SELECT 1
                FROM historial_tareas h2
                WHERE h2.op_no = t.titulo
                  AND COALESCE(h2.cierre_automatico, 0) = 1
            ) AS tiene_cierre_auto
        FROM tareas t
        WHERE t.estado = 'Finalizado'
          AND (t.asignado_a IS NULL OR TRIM(t.asignado_a) = '')
        ORDER BY t.fin DESC, t.id DESC
    """).fetchall()

    conn.close()

    return render_template(
        'historial.html',
        tareas=tareas,
    )

def _historial_central_base():
    return HISTORIAL_CENTRAL_DIR


def _safe_historial_relpath(relpath: str):
    rel = (relpath or "").replace("\\", "/").lstrip("/")
    if not rel:
        return None, None
    if ".." in rel.split("/"):
        return None, None

    base_abs = os.path.abspath(_historial_central_base())
    full_abs = os.path.abspath(os.path.join(base_abs, rel))
    if os.path.commonpath([base_abs, full_abs]) != base_abs:
        return None, None
    return rel, full_abs


def _listar_archivos_historial_central():
    base = _historial_central_base()
    if not os.path.isdir(base):
        os.makedirs(base, exist_ok=True)

    # Si el historial central está vacío, crear un respaldo inicial
    # para que la ventana tenga archivos desde la primera apertura.
    tiene_archivos = False
    for _root, _dirs, files in os.walk(base):
        if any(f and not f.startswith("_") for f in files):
            tiene_archivos = True
            break

    if not tiene_archivos:
        try:
            conn_boot = get_db_connection()
            respaldar_historial_en_central(conn_boot)
            conn_boot.close()
        except Exception:
            try:
                conn_boot.close()
            except Exception:
                pass

    out = []
    for root, _, files in os.walk(base):
        for nombre in files:
            if nombre.startswith("_"):
                continue
            full_path = os.path.join(root, nombre)
            if not os.path.isfile(full_path):
                continue

            relpath = os.path.relpath(full_path, base).replace("\\", "/")
            partes = relpath.split("/")
            bloque = partes[0] if partes else ""
            dia = partes[1] if len(partes) >= 3 else ""

            if bloque.startswith("Historial_admin_"):
                tipo = "admin"
                mes = bloque.replace("Historial_admin_", "", 1)
            elif bloque.startswith("Historial_usuario_"):
                tipo = "usuario"
                mes = bloque.replace("Historial_usuario_", "", 1)
            else:
                tipo = "otro"
                mes = ""

            try:
                st = os.stat(full_path)
                size_bytes = int(st.st_size)
                mtime_ts = float(st.st_mtime)
            except OSError:
                size_bytes = 0
                mtime_ts = 0.0

            out.append(
                {
                    "relpath": relpath,
                    "full_path": full_path,
                    "nombre": nombre,
                    "tipo": tipo,
                    "mes": mes,
                    "dia": dia,
                    "size_bytes": size_bytes,
                    "mtime_ts": mtime_ts,
                    "mtime": datetime.fromtimestamp(mtime_ts).strftime("%Y-%m-%d %H:%M:%S")
                    if mtime_ts
                    else "-",
                    "ext": os.path.splitext(nombre)[1].lower(),
                }
            )

    out.sort(key=lambda x: (x["mtime_ts"], x["relpath"]), reverse=True)
    return out


@admin_bp.route('/historial_archivos', methods=['GET'])
def historial_archivos():
    if not _solo_admin():
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("login"))

    tipo = (request.args.get("tipo") or "").strip().lower()
    mes = (request.args.get("mes") or "").strip()
    q = (request.args.get("q") or "").strip()
    archivo = (request.args.get("archivo") or "").strip().replace("\\", "/")

    rows_all = _listar_archivos_historial_central()
    meses = sorted({r["mes"] for r in rows_all if r["mes"]}, reverse=True)

    rows = rows_all
    if tipo in ("admin", "usuario"):
        rows = [r for r in rows if r["tipo"] == tipo]
    if mes:
        rows = [r for r in rows if r["mes"] == mes]
    if q:
        qq = q.lower()
        rows = [r for r in rows if qq in r["relpath"].lower()]

    total_filtrado = len(rows)
    limite_selector = 5
    rows_window = rows[:limite_selector]

    if not archivo and rows_window:
        archivo = rows_window[0]["relpath"]

    seleccionado = None
    selected_idx = None
    for idx, r in enumerate(rows_window, start=1):
        if r["relpath"] == archivo:
            seleccionado = r
            selected_idx = idx
            break

    preview_headers = []
    preview_rows = []
    preview_error = ""
    download_url = None

    if seleccionado:
        download_url = url_for("admin.historial_archivos_descargar", relpath=seleccionado["relpath"])
        if seleccionado["ext"] == ".csv":
            import csv

            try:
                with open(seleccionado["full_path"], "r", encoding="utf-8", errors="replace", newline="") as f:
                    reader = csv.reader(f)
                    preview_headers = next(reader, [])
                    for i, fila in enumerate(reader):
                        if i >= 30:
                            break
                        preview_rows.append(fila)
            except Exception as e:
                preview_error = str(e)

    return render_template(
        "admin/historial_archivos.html",
        rows=rows_window,
        total=total_filtrado,
        visible=len(rows_window),
        limite_selector=limite_selector,
        selected_idx=selected_idx or 0,
        meses=meses,
        tipo=tipo,
        mes=mes,
        q=q,
        archivo=archivo,
        seleccionado=seleccionado,
        download_url=download_url,
        preview_headers=preview_headers,
        preview_rows=preview_rows,
        preview_error=preview_error,
    )


@admin_bp.route('/historial_archivos/descargar/<path:relpath>', methods=['GET'])
def historial_archivos_descargar(relpath):
    if not _solo_admin():
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for("login"))

    rel, full_abs = _safe_historial_relpath(relpath)
    if not rel or not full_abs:
        abort(400)
    if not os.path.isfile(full_abs):
        abort(404)

    return send_from_directory(_historial_central_base(), rel, as_attachment=True)


# ==========================================================
# LIMPIAR HISTORIAL (DEMO) â€” SIN BORRAR USUARIOS NI OP
# ==========================================================
@admin_bp.route('/limpiar_historial', methods=['POST'])
def limpiar_historial():
    if session.get('clave_tecnica') != CLAVE_TECNICA:
        flash("⚠️ Acceso denegado. Entra primero por Modo Técnico.", "danger")
        return redirect(url_for('admin.panel'))

    conn = get_db_connection()
    try:
        resumen = limpiar_historial_con_respaldo(conn)
        flash("✅ Historial limpiado para demo. Usuarios y OP activas se mantienen intactos.", "success")
        flash(
            f"📁 Respaldo central creado. Admin: {resumen['admin_rows']} filas, Usuario: {resumen['usuario_rows']} filas.",
            "info",
        )
    except Exception as e:
        conn.rollback()
        flash(f"⚠️ Error al limpiar historial: {e}", "danger")
    finally:
        conn.close()

    return redirect(url_for('admin.panel'))


# ==========================================================
# MODO TÃ‰CNICO / BACKUPS
# ==========================================================
@admin_bp.route('/modo_tecnico', methods=['GET', 'POST'])
def modo_tecnico():
    import datetime, os
    if request.method == 'POST':
        clave = request.form.get('clave_tecnica', '').strip()
        if clave != CLAVE_TECNICA:
            flash("Clave técnica incorrecta.", "danger")
            return redirect(url_for('admin.panel'))
        session['clave_tecnica'] = CLAVE_TECNICA
    else:
        if session.get('clave_tecnica') != CLAVE_TECNICA:
            flash("Acceso denegado al modo técnico.", "danger")
            return redirect(url_for('admin.panel'))

    conn = get_db_connection()
    cur = conn.cursor()
    usuarios = cur.execute("SELECT * FROM usuarios ORDER BY id DESC").fetchall()
    conn.close()

    base_dir = os.path.dirname(os.path.abspath(__file__))          # .../admin
    app_dir = os.path.abspath(os.path.join(base_dir, ".."))        # .../HELP_APP
    backups_root = os.path.join(app_dir, "backups")

    backups = []
    for root, dirs, files in os.walk(backups_root):
        for d in dirs:
            if d.startswith("HELP_APP_BACKUP_"):
                full_path = os.path.join(root, d)
                fecha_str = d.replace("HELP_APP_BACKUP_", "")[:15]
                try:
                    fecha = datetime.datetime.strptime(fecha_str, "%Y%m%d_%H%M%S")
                    fecha_legible = fecha.strftime("%d/%m/%Y â€” %I:%M:%S %p")
                except:
                    fecha_legible = "Fecha desconocida"
                backups.append({"nombre": d, "ruta": full_path, "fecha": fecha_legible})

    backups = sorted(backups, key=lambda x: x["nombre"], reverse=True)
    return render_template('modo_tecnico.html', usuarios=usuarios, backups=backups)


@admin_bp.route('/reset_password_usuario/<int:user_id>', methods=['POST'])
def reset_password_usuario(user_id):
    from werkzeug.security import generate_password_hash

    clave = (request.form.get('clave_tecnica', '') or session.get('clave_tecnica', '')).strip()
    if clave != CLAVE_TECNICA:
        flash("Clave técnica incorrecta. No se realizó el cambio.", "danger")
        return redirect(url_for('admin.panel'))

    conn = get_db_connection()
    cur = conn.cursor()
    row = cur.execute("SELECT id, nombre FROM usuarios WHERE id=?", (user_id,)).fetchone()
    if not row:
        conn.close()
        flash("âš ï¸ Usuario no encontrado.", "warning")
        return redirect(url_for('admin.panel'))

    nueva = '1234'
    hashed = generate_password_hash(nueva, method='scrypt')

    # âœ… Reset + obligar cambio en el prÃ³ximo login
    cur.execute(
        "UPDATE usuarios SET contrasena=?, must_change_password=1 WHERE id=?",
        (hashed, user_id)
    )

    conn.commit()
    conn.close()

    flash(f"âœ… ContraseÃ±a de '{row['nombre']}' reseteada a '1234'. El usuario deberÃ¡ crear una nueva al ingresar.", "success")
    return redirect(url_for('admin.panel'))



@admin_bp.route('/cambiar_password_usuario/<int:user_id>', methods=['POST'])
def cambiar_password_usuario(user_id):
    from werkzeug.security import generate_password_hash

    # Misma protecciÃ³n de clave tÃ©cnica
    clave = (request.form.get('clave_tecnica', '') or session.get('clave_tecnica', '')).strip()
    if clave != CLAVE_TECNICA:
        flash("Clave técnica incorrecta. No se realizó el cambio.", "danger")
        return redirect(url_for('admin.panel'))

    nueva = (request.form.get('nueva_contrasena', '') or '').strip()
    if not nueva:
        flash("âš ï¸ Debes escribir una nueva contraseÃ±a.", "warning")
        return redirect(url_for('admin.panel'))

    conn = get_db_connection()
    cur = conn.cursor()
    row = cur.execute("SELECT id, nombre FROM usuarios WHERE id=?", (user_id,)).fetchone()
    if not row:
        conn.close()
        flash("âš ï¸ Usuario no encontrado.", "warning")
        return redirect(url_for('admin.panel'))

    hashed = generate_password_hash(nueva, method='scrypt')
    cur.execute("UPDATE usuarios SET contrasena=? WHERE id=?", (hashed, user_id))
    conn.commit()
    conn.close()

    nombre = row[1] if not hasattr(row, "keys") else row["nombre"]
    flash(f"âœ… ContraseÃ±a de '{nombre}' cambiada correctamente.", "success")
    return redirect(url_for('admin.panel'))



@admin_bp.route('/abrir_backup/<nombre>')
def abrir_backup(nombre):
    import os

    if session.get('clave_tecnica') != CLAVE_TECNICA:
        flash("Acceso denegado al modo técnico.", "danger")
        return redirect(url_for('admin.panel'))

    base_dir = os.path.dirname(os.path.abspath(__file__))   # .../admin
    app_dir = os.path.abspath(os.path.join(base_dir, "..")) # .../HELP_APP
    backups_root = os.path.join(app_dir, "backups")

    for root, dirs, files in os.walk(backups_root):
        for d in dirs:
            if d == nombre:
                ruta = os.path.join(root, d)
                try:
                    os.startfile(ruta)
                    flash(f"ðŸ“‚ Carpeta abierta: {nombre}", "success")
                except Exception as e:
                    flash(f"âš ï¸ No se pudo abrir la carpeta: {e}", "danger")
                return redirect(url_for('admin.panel'))

    flash("âš ï¸ No se encontrÃ³ el backup solicitado.", "warning")
    return redirect(url_for('admin.panel'))


@admin_bp.route('/exportar_excel', methods=['POST'])
def exportar_excel():
    if not _solo_admin():
        flash("Acceso no autorizado.", "danger")
        return redirect(url_for('login'))

    conn = get_db_connection()
    exportar_excel_automatico(conn)
    conn.close()

    flash("Se generÃ³ la exportaciÃ³n en Excel (.xlsx).", "success")
    return redirect(url_for('admin.panel'))

# ==========================================
# ðŸ“ Carpeta diaria (Web) - PythonAnywhere
# ==========================================
import os
import csv
from io import BytesIO
from datetime import datetime
from flask import render_template, send_file, send_from_directory, flash, redirect, url_for, abort
from openpyxl import Workbook, load_workbook


def _carpeta_diaria_base():
    # Usar exactamente la misma ruta base del registro diario para evitar desalineaciÃ³n.
    return CARPETA_TAREAS_DIARIAS


def _carpeta_por_fecha(fecha=None):
    # fecha: "YYYY_MM_DD" o None = hoy
    if not fecha:
        fecha = datetime.now().strftime("%Y_%m_%d")
    return os.path.join(_carpeta_diaria_base(), fecha)


def _ruta_csv(fecha=None):
    return os.path.join(_carpeta_por_fecha(fecha), "tareas.csv")


def _ruta_xlsx(fecha=None):
    return os.path.join(_carpeta_por_fecha(fecha), "tareas.xlsx")


def _leer_historial_diario(fecha=None):
    """
    Lee historial diario desde CSV (si existe) o desde XLSX como fallback.
    Devuelve: (headers, rows, ruta_origen).
    """
    ruta_csv = _ruta_csv(fecha)
    if os.path.exists(ruta_csv):
        with open(ruta_csv, "r", encoding="utf-8", newline="") as f:
            reader = csv.reader(f)
            try:
                headers = next(reader)
            except StopIteration:
                return [], [], ruta_csv
            rows = list(reader)
        return headers, rows, ruta_csv

    ruta_xlsx = _ruta_xlsx(fecha)
    if os.path.exists(ruta_xlsx):
        wb = load_workbook(ruta_xlsx, data_only=True, read_only=True)
        try:
            ws = wb["Tareas"] if "Tareas" in wb.sheetnames else wb.active
            filas = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()

        if not filas:
            return [], [], ruta_xlsx

        headers = ["" if v is None else str(v).strip() for v in filas[0]]
        total_cols = len(headers)
        rows = []
        for fila in filas[1:]:
            rr = list(fila[:total_cols])
            if len(rr) < total_cols:
                rr.extend([None] * (total_cols - len(rr)))
            rr = ["" if v is None else str(v).strip() for v in rr]
            if any(c != "" for c in rr):
                rows.append(rr)
        return headers, rows, ruta_xlsx

    return [], [], ruta_xlsx


def _leer_checklist_diario_desde_xlsx(fecha=None):
    """
    Lee hoja 'checklist' desde el XLSX diario, si existe.
    Devuelve: (headers, rows).
    """
    ruta_xlsx = _ruta_xlsx(fecha)
    if not os.path.exists(ruta_xlsx):
        return [], []

    wb = load_workbook(ruta_xlsx, data_only=True, read_only=True)
    try:
        if "checklist" not in wb.sheetnames:
            return [], []
        ws = wb["checklist"]
        filas = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not filas:
        return [], []

    headers = ["" if v is None else str(v).strip() for v in filas[0]]
    total_cols = len(headers)
    rows = []
    for fila in filas[1:]:
        rr = list(fila[:total_cols])
        if len(rr) < total_cols:
            rr.extend([None] * (total_cols - len(rr)))
        rr = ["" if v is None else str(v).strip() for v in rr]
        if any(c != "" for c in rr):
            rows.append(rr)
    return headers, rows


# ==========================================================
# âœ… Helpers: formateo seguro HH:MM (SIN tocar el CSV en disco)
# ==========================================================
def _fmt_hhmm(value):
    """
    Convierte 'YYYY-MM-DD HH:MM:SS' o 'HH:MM:SS' o 'HH:MM' -> 'HH:MM'.
    Si no se puede convertir, devuelve el valor original.
    """
    if value is None:
        return value
    s = str(value).strip()
    if not s:
        return s

    # si viene con fecha, deja solo la hora
    if " " in s:
        s = s.split(" ")[-1]
    if "T" in s:
        s = s.split("T")[-1]

    parts = s.split(":")
    if len(parts) >= 2:
        try:
            h = int(parts[0])
            m = int(parts[1])
            return f"{h:02d}:{m:02d}"
        except Exception:
            return s
    return s


def _detectar_idxs_hora(headers):
    """
    Detecta de forma flexible columnas de inicio/fin/tiempo.
    Soporta: inicio, hora_inicio, fin, hora_fin, hora_finalizacion,
             tiempo_tarea, tiempo_total, duracion, duraciÃ³n...
    """
    lower = [str(h).strip().lower() for h in headers]
    idx_inicio = None
    idx_fin = None
    idx_tiempo = None

    for i, h in enumerate(lower):
        # inicio: contiene "inicio" y ademÃ¡s suena a hora
        if idx_inicio is None and ("inicio" in h) and (("hora" in h) or (h == "inicio") or ("time" in h)):
            idx_inicio = i

        # fin: fin/final/finalizacion y suena a hora
        if idx_fin is None and (("fin" in h) or ("final" in h) or ("finalizacion" in h) or ("finalizaciÃ³n" in h)) and (("hora" in h) or (h == "fin") or ("time" in h)):
            idx_fin = i

        # tiempo: tiempo/duracion
        if idx_tiempo is None and (("tiempo" in h) or ("duracion" in h) or ("duraciÃ³n" in h)):
            idx_tiempo = i

    return idx_inicio, idx_fin, idx_tiempo


def _aplicar_hhmm_a_tabla(headers, rows):
    """
    Aplica HH:MM solo a columnas detectadas (inicio/fin/tiempo).
    Devuelve (headers, rows_formateadas).
    """
    if not headers or not rows:
        return headers, rows

    idx_inicio, idx_fin, idx_tiempo = _detectar_idxs_hora(headers)
    idxs = [i for i in (idx_inicio, idx_fin, idx_tiempo) if i is not None]
    if not idxs:
        return headers, rows

    out = []
    for r in rows:
        rr = list(r)
        for i in idxs:
            if i < len(rr):
                rr[i] = _fmt_hhmm(rr[i])
        out.append(rr)

    return headers, out


@admin_bp.route("/carpeta_diaria", defaults={"fecha": None})
@admin_bp.route("/carpeta_diaria/<fecha>")
def carpeta_diaria(fecha):
    carpeta = _carpeta_por_fecha(fecha)
    if not os.path.isdir(carpeta):
        archivos = []
    else:
        archivos = sorted(os.listdir(carpeta))

    return render_template("admin/carpeta_diaria.html", carpeta=carpeta, fecha=fecha, archivos=archivos)

@admin_bp.route("/carpeta_diaria/descargar/<path:nombre>")
def carpeta_diaria_descargar_hoy(nombre):
    carpeta = _carpeta_por_fecha(None)  # hoy
    if not os.path.isdir(carpeta):
        abort(404)
    return send_from_directory(carpeta, nombre, as_attachment=True)



@admin_bp.route("/carpeta_diaria/<fecha>/descargar/<path:nombre>")
def carpeta_diaria_descargar(fecha, nombre):
    carpeta = _carpeta_por_fecha(fecha)
    if not os.path.isdir(carpeta):
        abort(404)
    return send_from_directory(carpeta, nombre, as_attachment=True)


# ==========================================
# ðŸ“‹ Historial diario (HOY) - Tabla y descargas
# ==========================================
@admin_bp.route("/historial_diario", methods=["GET"])
def historial_diario():
    headers, rows, ruta = _leer_historial_diario(None)
    if not headers and not rows:
        flash(f"No existe el historial de hoy todavia: {ruta}", "warning")
        return render_template("admin/historial_diario.html", headers=[], rows=[], ruta=ruta)

    # âœ… Formateo SOLO para mostrar en web (NO modifica el CSV en disco)
    headers, rows = _aplicar_hhmm_a_tabla(headers, rows)

    return render_template("admin/historial_diario.html", headers=headers, rows=rows, ruta=ruta)


@admin_bp.route("/historial_diario.csv", methods=["GET"])
def descargar_historial_diario_csv():
    headers, rows, _ruta = _leer_historial_diario(None)
    if not headers and not rows:
        flash("No hay historial de hoy para descargar.", "warning")
        return redirect(url_for("admin.historial_diario"))

    # âœ… Formateo SOLO para el CSV descargado (NO modifica el CSV en disco)
    headers, rows = _aplicar_hhmm_a_tabla(headers, rows)

    # Generar CSV en memoria
    import io
    text_buf = io.StringIO()
    w = csv.writer(text_buf, lineterminator="\n")
    w.writerow(headers)
    w.writerows(rows)

    output = BytesIO(text_buf.getvalue().encode("utf-8"))
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"tareas_{datetime.now().strftime('%Y_%m_%d')}.csv",
        mimetype="text/csv"
    )


@admin_bp.route("/historial_diario.xlsx", methods=["GET"])
def descargar_historial_diario_xlsx():
    headers, rows, _ruta = _leer_historial_diario(None)
    if not headers and not rows:
        flash("No hay historial de hoy para generar Excel.", "warning")
        return redirect(url_for("admin.historial_diario"))

    # âœ… Formateo SOLO para el XLSX (NO modifica el CSV en disco)
    headers, rows = _aplicar_hhmm_a_tabla(headers, rows)
    data = [headers] + rows

    wb = Workbook()
    ws = wb.active
    ws.title = "tareas"

    for row in data:
        ws.append(row)

    # ✅ Incluir hoja checklist cuando exista en el XLSX diario original.
    chk_headers, chk_rows = _leer_checklist_diario_desde_xlsx(None)
    if chk_headers:
        ws_chk = wb.create_sheet(title="checklist")
        ws_chk.append(chk_headers)
        for row in chk_rows:
            ws_chk.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"tareas_{datetime.now().strftime('%Y_%m_%d')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


